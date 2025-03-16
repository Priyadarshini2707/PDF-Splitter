from flask import Flask, request, send_from_directory, jsonify
from flask_cors import CORS
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import load_workbook

# Initialize Flask app
app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER

# Ensure directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Function to read Excel and write its content to a PDF
def excel_to_pdf(excel_path, pdf_path, sheet_names):
    workbook = load_workbook(excel_path)
    canvas_obj = canvas.Canvas(pdf_path, pagesize=A4)

    width, height = A4
    x_margin, y_margin = 50, 50  # Margins for content
    line_spacing = 20  # Space between lines
    y = height - y_margin  # Start drawing from the top margin

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        sheet = workbook[sheet_name]

        # Draw sheet name at the top
        canvas_obj.setFont("Helvetica-Bold", 12)
        canvas_obj.drawString(x_margin, y, f"Sheet: {sheet_name}")
        y -= line_spacing

        # Write data from rows
        canvas_obj.setFont("Helvetica", 10)
        for row in sheet.iter_rows(values_only=True):
            line = " | ".join(str(cell) if cell is not None else "" for cell in row)
            canvas_obj.drawString(x_margin, y, line)
            y -= line_spacing

            # Start a new page if the content exceeds the page
            if y < y_margin:
                canvas_obj.showPage()
                y = height - y_margin
                canvas_obj.setFont("Helvetica-Bold", 12)
                canvas_obj.drawString(x_margin, y, f"Sheet: {sheet_name}")
                y -= line_spacing

    canvas_obj.save()

# Endpoint to convert Excel to PDF
@app.route('/convert/html', methods=['POST'])
def convert_excel_to_pdf_html_endpoint():
    try:
        file = request.files.get('file')
        sheet_names = request.form.get('sheetNames').split(',')

        if not file or not sheet_names:
            return jsonify({"error": "Missing file or sheet names"}), 400

        # Save uploaded Excel file
        input_excel_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(input_excel_path)

        # Generate output PDF
        output_pdf_path = os.path.join(app.config['OUTPUT_FOLDER'], 'output.pdf')
        excel_to_pdf(input_excel_path, output_pdf_path, sheet_names)

        # Serve the generated PDF
        return send_from_directory(app.config['OUTPUT_FOLDER'], 'output.pdf', as_attachment=True)

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True)
